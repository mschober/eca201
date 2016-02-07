import csv
from openpyxl import Workbook
from openpyxl.cell import get_column_letter

class XlsConv:

    def __init__(self, src_file, tgt_file, title="output.xlsx"):
        self.src_file = src_file
        self.tgt_file = tgt_file
        self.wb = Workbook()
        self.ws = self.wb.worksheets[0]
        self.ws.title = title
        self.reader = self.__make_reader(src_file)

    def __make_reader(self, src_file):
        src_file_handle = open(r'{0}'.format(src_file), 'rU')
        return csv.reader(src_file_handle)


    def __save_file(self, tgt_file, wb):
        dest_filename = r"{0}".format(tgt_file)
        wb.save(filename = dest_filename)


    def convert(self):
        print "The source file is %s" % self.src_file
        print "The target file is %s" % self.tgt_file

        print "Begin converting to xlsx..."
        for row_index, row in enumerate(self.reader):
            for column_index, cell in enumerate(row):
                column_letter = get_column_letter((column_index + 1))
                self.ws.cell('%s%s'%(column_letter, (row_index + 1))).value = cell

        print "Saving the converted file..."
        self.__save_file(self.tgt_file, self.wb)
